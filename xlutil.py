import openpyxl

#save xlsx file location to "file" variable
#save sheet name to "sheet" variable

#use this function to take the count of total rows

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

#use this function to take the count of total columns

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

#use this function to read data

def readData(file,sheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnnum).value

#use this function to write data

def writeData(file,sheetName,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum,column=columnnum).value=data
    workbook.save(file)
